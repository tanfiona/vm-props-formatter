import copy
import openpyxl
import pandas as pd
import numpy as np
import xlrd
#import glob
#import os
#from .utils.file_organizer import check_create_directory


class VMPropsManager(object):
    """"""
    __data_parser = None
    __parameters = None
    __defaults = {
        "shape": {
            "title_row": 4,
            "title_col": 1,
            "main_header_row": 7,
            "number_of_header_rows": 1,
            "props_header_start_col": 5,
            "props_header_tally_first": 5,
            "no_summary_table_rows": 3,
            "summary_table_sum_row": 1,
            "props_header_end_col": -1
        },
        "names": {
            "sheet_name": "",
            "country_col": "COUNTRY NAME",
            "store_col": "STORE NAME",
            "storesap_col": "",
            "main_cols": ["COUNTRY NAME"],
            "entity_list": ["CKS", "CKI", "CKC"],
            "drop_rows_with": {
                "NO": ["Total:"]
            }
        }
    }

    def __init__(self, parameters=None):
        """
        Constructor that copies the parameters

        Parameters
        ----------
        parameters : dict
            Dictionary of parameters

        Returns
        -------
        None
        """
        # Set defined parameter values if present
        if parameters is not None:
            self.__parameters = parameters
        else:
            self.__parameters = copy.deepcopy(self.__defaults)

    def update_parameters(self, parameters):
        """
        Update the parameters

        Parameters
        ----------
        parameters : dict
            Dictionary of parameters

        Returns
        -------
        None
        """
        self.__parameters.update(parameters)

    def get_default_parameters(self):
        """
        Get the default parameters

        Parameters
        ----------

        Returns
        -------
        parameters : dict
            Default parameters
        """
        return copy.deepcopy(self.__defaults)

    def get_entity(self, file_name):
        entity = None
        for i in self.__parameters['names']['entity_list']:
            if i in file_name:
                entity = i
        return entity

    def load_dataset(self, file_path, file_name, file_only=False, sheet_name=None, header=None, import_merged=False):
        """
        Loads excel files into data and colour information
        
        Parameters
        ----------
        file_path : byte
        file_name : str
        file_only : bool
        sheet_name : str
        header : int
        import_merged : bool

        Returns
        -------
        data: pandas DataFrame
        sh : object

        """
        # load sheet name from parameters
        if sheet_name is None:
            sheet_name = self.__parameters['names']['sheet_name']
        # open file
        book = pd.ExcelFile(file_path).book
        # take first, if sheet_name is blank/ not specified
        if sheet_name == '':
            # check all sheet names in book
            sheets = book.sheets()
            visible_sheets = []
            for sheet in sheets:
                if sheet.visibility == 0:  # sheet is visible
                    visible_sheets.append(sheet.name)
            # take first visible sheet
            sheet_name = str(visible_sheets[0])
            print('[Status] Sheet name not specified. Took sheet by loc: ', sheet_name)
        print('[Status] Loading File: "%s";' % file_name, '"Loading Sheet: "%s"' % sheet_name)
        if not import_merged:
            data = pd.read_excel(file_path, sheet_name, index_col=None, header=header)
        else:
            # read file data by sheet_name
            sheet = book.sheet_by_name(sheet_name)
            # get and overwrite merged cells
            data = []
            for row_index in range(sheet.nrows):
                row = []
                for col_index in range(sheet.ncols):
                    valor = sheet.cell(row_index, col_index).value
                    if valor == '':
                        for crange in sheet.merged_cells:
                            rlo, rhi, clo, chi = crange
                            if rlo <= row_index < rhi and clo <= col_index < chi:
                                valor = sheet.cell(rlo, clo).value
                                break
                    row.append(valor)
                data.append(row)
            data = pd.DataFrame(data)
        # remove leading and trailing whitespaces in cells
        data = data.applymap(lambda x: str(x).strip())
        # remove None types in different formats
        data = data.replace(["NA", "NONE", "NAN", "NULL", ""], np.nan)
        # read colours info
        if not file_only:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sh = wb[sheet_name]
            return data, sh, sheet_name
        else:
            return data

    def rename_duplicate_column_names(self, df):
        # df is the dataframe that you want to rename duplicated columns
        cols = pd.Series(df.columns)
        for dup in cols[cols.duplicated()].unique():
            cols[cols[cols == dup].index.values.tolist()] = [dup + '_' + str(i) if i != 0 else dup for i in
                                                             range(sum(cols == dup))]

        # rename the columns with the cols list.
        df.columns = cols
        return df

    def get_main_data(self, data, skiprows = None, headerrows= None):
        """
        Shortens and re-header data

        Parameters
        ----------
        data: pandas.DataFrame
            Dataframe to run the operation on
        skiprows: int
            Number of rows to skip to first header row
            E.g. skiprows = 7 if the main data header is on Row '8'
        headerrows: int
            Number of rows that make up headers
            E.g. 2 if both window props + set A both are header rows

        Returns
        -------

        """
        # load parameters if not specified
        if skiprows is None:
            skiprows = self.__parameters['shape']['main_header_row']
        if headerrows is None:
            headerrows = self.__parameters['shape']['number_of_header_rows']
        # run analysis
        new_header = data.iloc[skiprows+headerrows-1]
        data = data.iloc[skiprows+headerrows:,]
        data.columns = new_header
        return data

    def dropna_rows_cols(self, data):
        """

        Parameters
        ----------
        data

        Returns
        -------

        """
        # drop rows that have na
        data = data.dropna(how='all')
        # get idx of last col with value
        last_idx = max(np.where(~data.columns.isna())[0]) + 1
        # drop cols that have na AFTER main frame
        data = data.iloc[:, :last_idx]
        return data

    def get_index_to_split_tables2(self, data_1, main_cols=None):
        # load parameters if not specified
        if main_cols is None:
            main_cols = self.__parameters['names']['main_cols']
        return [data_1[data_1[main_cols].isna().all(axis=1)].index[0]]

    def get_index_to_split_tables(self, main_data, skipcols=None, tallycols=None):
        """ 
        Function to find number of split tables based on first row header name
        
        Parameters
        ----------
        skipcols
        tallycols

        Returns
        -------

        """
        # load parameters if not specified
        if skipcols is None:
            skipcols = self.__parameters['shape']['props_header_start_col']
        if tallycols is None:
            tallycols = self.__parameters['shape']['props_header_tally_first']
        # run analysis
        col_name_list = []
        # get row indexes which tally with colnames
        sub_data = main_data.iloc[:, skipcols:skipcols + tallycols]
        for i in list(sub_data.index):
            if list(sub_data.loc[i].values) == list(sub_data.loc[i].index):
                col_name_list.append(i)
        return col_name_list

    def get_split_data(self, data, col_name_list):
        """
        Split data by col_name_list

        Parameters
        ----------
        data
        col_name_list

        Returns
        -------

        """
        # run analysis
        if len(col_name_list) == 1:
            data_1 = data.loc[:col_name_list[0]-1,]
            data_2 = data.loc[col_name_list[0]:,]
            return data_1, data_2
        elif len(col_name_list) == 2:
            data_1 = data.loc[:col_name_list[0]-1,]
            data_2 = data.loc[col_name_list[0]:col_name_list[1]-1,]
            data_3 = data.loc[col_name_list[1]:,]
            return data_1, data_2, data_3

    def clean_main_data(self, data, country_col=None, drop_rows_with=None):
        """
        
        Parameters
        ----------
        data
        country_col

        Returns
        -------

        """
        # load parameters if not specified
        if country_col is None:
            country_col = self.__parameters['names']['country_col']
        if drop_rows_with is None:
            drop_rows_with = self.__parameters['names']['drop_rows_with']
        # run analysis
        # replace duplicate columns
        data = self.rename_duplicate_column_names(data)
        return data

    def shorten_table_w_max_rows(self, data, max_rows=None):
        """
        
        Parameters
        ----------
        data: pandas.DataFrame
        max_rows: int

        Returns
        -------

        """
        # load parameters if not specified
        if max_rows is None:
            max_rows = self.__parameters['shape']['no_summary_table_rows']
        # run analysis
        return data.iloc[:max_rows,]

    def format_main_data(self, data, country_col=None, drop_rows_with=None, skipcols_front=None, skipcols_end=None):
        """
        
        Parameters
        ----------
        data

        Returns
        -------

        """
        # load parameters if not specified
        if country_col is None:
            country_col = self.__parameters['names']['country_col']
        if drop_rows_with is None:
            drop_rows_with = self.__parameters['names']['drop_rows_with']
        if skipcols_front is None:
            skipcols_front = self.__parameters['shape']['props_header_start_col']
        if skipcols_end is None:
            skipcols_end = self.__parameters['shape']['props_header_end_col']
        # start analysis
        # drop 'duplicate' rows where any of the Total word exists
        for cell_value in drop_rows_with:
            data = data[~pd.DataFrame(data == cell_value).any(axis='columns')]

        # format country
        country_col_2_ind = data.columns.get_loc(country_col) + 1
        data[country_col] = data.iloc[:, country_col_2_ind].fillna(data[country_col])

        # fill na with above country (some not merged properly)
        data.loc[:, country_col] = data.loc[:, country_col].ffill()

        # fillna
        df = data.fillna(0)

        # clear white space cells
        df = df.applymap(lambda x: 0 if str(x).isspace() else x)

        # format prop columns into numeric
        for col_idx in range(skipcols_front, df.shape[1]+skipcols_end):
            df.iloc[:, col_idx] = pd.to_numeric(df.iloc[:, col_idx], errors='coerce')

        # add row sum
        df.loc[max(df.index) + 1] = df.sum(numeric_only=True)

        # fillna
        df = df.fillna(0)

        return df

    def main_and_summary_checker(self, df, summary, sum_row=None, skipcols_front=None, skipcols_end=None):
        # load parameters if not specified
        if sum_row is None:
            sum_row = self.__parameters['shape']['summary_table_sum_row']
        if skipcols_front is None:
            skipcols_front = self.__parameters['shape']['props_header_start_col']
        if skipcols_end is None:
            skipcols_end = self.__parameters['shape']['props_header_end_col']
        # run analysis
        checker = pd.DataFrame()
        checker['VM PROPS'] = list(df.iloc[:, skipcols_front:skipcols_end].columns)
        checker['main'] = pd.to_numeric(list(df.iloc[-1, skipcols_front:skipcols_end].fillna(0)))
        checker['summary'] = pd.to_numeric(list(summary.iloc[sum_row, skipcols_front:skipcols_front+len(checker)].fillna(0)))
        checker['checks'] = list(checker['main'] == checker['summary'])
        return checker

    def get_excel_col_from_int(self, n):
        string = ""
        n = int(n)
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            string = chr(65 + remainder) + string
        return string
    
    def main_table_to_so_converter(self, df, main_cols=None, skipcols_front=None, skipcols_end=None):
        # load parameters if not specified
        if main_cols is None:
            main_cols = self.__parameters['names']['main_cols']
        if skipcols_front is None:
            skipcols_front = self.__parameters['shape']['props_header_start_col']
        if skipcols_end is None:
            skipcols_end = self.__parameters['shape']['props_header_end_col']
        # run analysis
        # get prop column name list from column location
        props_column_names = list(df.iloc[:, skipcols_front:skipcols_end].columns)
        # melt all props into country
        df['XRow'] = df.index + 1
        so_table = pd.melt(df[main_cols + ['XRow'] + props_column_names], id_vars=main_cols + ['XRow'],
                           value_vars=props_column_names)
        # rename column names to fixed format
        so_table.columns = main_cols + ['XRow', 'VM PROPS', 'Qty']
        # order columns
        so_table = so_table[main_cols + ['VM PROPS', 'Qty', 'XRow']]
        # drop props with no qty
        so_table['Qty'] = pd.to_numeric(so_table['Qty'], errors='coerce')
        so_table = so_table[so_table['Qty'] > 0]
        # calculate cell location (of original excel)
        max_rows = len(df)
        so_table['XCol'] = [skipcols_front + 1 + (i // max_rows) for i in so_table.index]  # quotient
        so_table['XCol'] = so_table['XCol'].apply(lambda x: self.get_excel_col_from_int(x))
        so_table['XCell'] = so_table['XCol'] + so_table['XRow'].astype('str')
        # rename total rows
        for col in main_cols:
            so_table.loc[so_table['XRow'] % (max(df.index) + 1) == 0, col] = 'TOTAL'
        # return so format table
        return so_table.reset_index(drop='True')

    def get_cell_colour_col(self, so_table, original_sheet):
        # run analysis
        so_table['Cell_Colour'] = so_table['XCell'].apply(lambda x: str(original_sheet[x].fill.start_color.index))
        so_table['Cell_Colour'] = so_table['Cell_Colour'].apply(lambda x: '00000000' if x == '0' else x)
        # rename total rows
        for col in ['XRow', 'XCol', 'XCell', 'Cell_Colour']:
            so_table.loc[so_table.iloc[:, 0] == 'TOTAL', col] = ''
        return so_table

    def get_file_name(self, sheet_name=None):
        # load parameters if not specified
        if sheet_name is None:
            sheet_name = self.__parameters['names']['sheet_name'] # default settings
        # run analysis
        if sheet_name is not None:
            report_filename = str(sheet_name) + ' VM Props Analysis Report.xlsx'
        else:
            report_filename = 'VM Props Analysis Report.xlsx'
        return report_filename